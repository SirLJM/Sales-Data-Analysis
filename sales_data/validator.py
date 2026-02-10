from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

from utils.logging_config import get_logger

logger = get_logger("validator")

DFRAME_EMPTY = "DataFrame is empty"


class DataValidator:
    SALES_COLUMNS = {"order_id", "data", "sku", "ilosc", "cena", "razem"}
    STOCK_COLUMNS = {
        "sku",
        "nazwa",
        "cena_netto",
        "cena_brutto",
        "stock",
        "available_stock",
        "aktywny",
    }
    FORECAST_COLUMNS = {"data", "sku", "forecast"}
    MODEL_METADATA_COLUMNS = {
        "Model",
        "SZWALNIA GŁÓWNA",
        "SZWALNIA DRUGA",
        "RODZAJ MATERIAŁU",
        "GRAMATURA",
    }
    CATEGORY_COLUMNS = {
        "Model",
        "Grupa",
        "Podgrupa",
        "Kategoria",
        "Nazwa"
    }
    BOM_COLUMNS = {"Model", "u Martyny nazwa", "norma mb/szt"}
    MATERIAL_CATALOG_COLUMNS = {"OPIS Z KARTY PRODUKTU", "RODZAJ MATERIAŁU", "DOSTAWCA"}

    @staticmethod
    def validate_sales_data(df: pd.DataFrame) -> tuple[bool, list[str]]:
        errors = []

        missing_columns = DataValidator.SALES_COLUMNS - set(df.columns)
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")

        if len(df) == 0:
            errors.append(DFRAME_EMPTY)

        if "ilosc" in df.columns and not pd.api.types.is_numeric_dtype(df["ilosc"]):
            errors.append("Column 'ilosc' must be numeric")

        if "cena" in df.columns and not pd.api.types.is_numeric_dtype(df["cena"]):
            errors.append("Column 'cena' must be numeric")

        if "razem" in df.columns and not pd.api.types.is_numeric_dtype(df["razem"]):
            errors.append("Column 'razem' must be numeric")

        is_valid = len(errors) == 0
        return is_valid, errors

    @staticmethod
    def validate_stock_data(df: pd.DataFrame) -> tuple[bool, list[str]]:
        errors = []

        missing_columns = DataValidator.STOCK_COLUMNS - set(df.columns)
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")

        if len(df) == 0:
            errors.append(DFRAME_EMPTY)

        if "stock" in df.columns and not pd.api.types.is_numeric_dtype(df["stock"]):
            errors.append("Column 'stock' must be numeric")

        if "available_stock" in df.columns and not pd.api.types.is_numeric_dtype(
                df["available_stock"]
        ):
            errors.append("Column 'available_stock' must be numeric")

        if "aktywny" in df.columns and not pd.api.types.is_numeric_dtype(df["aktywny"]):
            errors.append("Column 'aktywny' must be numeric (0 or 1)")

        is_valid = len(errors) == 0
        return is_valid, errors

    @staticmethod
    def validate_forecast_data(df: pd.DataFrame) -> tuple[bool, list[str]]:
        errors = []

        missing_columns = DataValidator.FORECAST_COLUMNS - set(df.columns)
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")

        if len(df) == 0:
            errors.append(DFRAME_EMPTY)

        if "forecast" in df.columns and not pd.api.types.is_numeric_dtype(df["forecast"]):
            errors.append("Column 'forecast' must be numeric")

        is_valid = len(errors) == 0
        return is_valid, errors

    @staticmethod
    def validate_model_metadata(df: pd.DataFrame) -> tuple[bool, list[str]]:
        errors = []

        missing_columns = DataValidator.MODEL_METADATA_COLUMNS - set(df.columns)
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")

        if len(df) == 0:
            errors.append(DFRAME_EMPTY)

        if "Model" not in df.columns:
            errors.append("Column 'Model' is required")

        is_valid = len(errors) == 0
        return is_valid, errors

    @staticmethod
    def validate_category_data(df: pd.DataFrame) -> bool:
        if df is None or df.empty:
            raise ValueError("Category data is empty")

        missing_cols = DataValidator.CATEGORY_COLUMNS - set(df.columns)
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")

        null_models = df["Model"].isna().sum()
        if null_models > 0:
            raise ValueError(f"Found {null_models} rows with null Model codes")

        invalid_models = df[df["Model"].str.len() != 5]
        if not invalid_models.empty:
            raise ValueError(f"Found {len(invalid_models)} models not 5 characters")

        for col in ["Grupa", "Podgrupa", "Kategoria"]:
            null_count = df[col].isna().sum()
            if null_count > 0:
                raise ValueError(f"Found {null_count} null values in {col}")

        return True

    @staticmethod
    def find_sheet_by_columns(
            file_path: Path, required_columns: set, required_column: str | None = None
    ) -> str | None:
        try:
            excel_file = pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                    if not required_columns.issubset(set(df.columns)):
                        continue
                    if required_column is None or required_column in df.columns:
                        return str(sheet_name)
                except (ValueError, KeyError):
                    continue
            return None
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {file_path}")
        except pd.errors.EmptyDataError:
            raise ValueError(f"File is empty: {file_path}")
        except PermissionError:
            raise PermissionError(f"Permission denied: {file_path}")
        except ValueError as e:
            raise ValueError(f"Error parsing file {file_path}: {e}") from e

    @staticmethod
    def find_sales_sheet(file_path: Path) -> str | None:
        return DataValidator.find_sheet_by_columns(
            file_path, DataValidator.SALES_COLUMNS, "data"
        )

    @staticmethod
    def find_stock_sheet(file_path: Path) -> str | None:
        return DataValidator.find_sheet_by_columns(
            file_path, DataValidator.STOCK_COLUMNS, "stock"
        )

    @staticmethod
    def find_forecast_sheet(file_path: Path) -> str | None:
        return DataValidator.find_sheet_by_columns(
            file_path, DataValidator.FORECAST_COLUMNS, "data"
        )

    @staticmethod
    def find_model_metadata_sheet(file_path: Path) -> str | None:
        return DataValidator.find_sheet_by_columns(
            file_path, DataValidator.MODEL_METADATA_COLUMNS, None
        )

    @staticmethod
    def find_bom_sheet(file_path: Path) -> str | None:
        return DataValidator.find_sheet_by_columns(
            file_path, DataValidator.BOM_COLUMNS, None
        )

    @staticmethod
    def find_material_catalog_sheet(file_path: Path) -> str | None:
        return DataValidator.find_sheet_by_columns(
            file_path, DataValidator.MATERIAL_CATALOG_COLUMNS, None
        )

    @staticmethod
    def validate_bom_data(df: pd.DataFrame) -> tuple[bool, list[str]]:
        errors = []

        missing_columns = DataValidator.BOM_COLUMNS - set(df.columns)
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")

        if len(df) == 0:
            errors.append(DFRAME_EMPTY)

        if "Model" not in df.columns:
            errors.append("Column 'Model' is required")

        is_valid = len(errors) == 0
        return is_valid, errors

    @staticmethod
    def validate_material_catalog(df: pd.DataFrame) -> tuple[bool, list[str]]:
        errors = []

        missing_columns = DataValidator.MATERIAL_CATALOG_COLUMNS - set(df.columns)
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")

        if len(df) == 0:
            errors.append(DFRAME_EMPTY)

        is_valid = len(errors) == 0
        return is_valid, errors

    @staticmethod
    def find_category_sheet(file_path: Path) -> str | None:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            if "Kategorie" in wb.sheetnames:
                sheet = wb["Kategorie"]
                headers = {str(cell.value) for cell in sheet[1] if cell.value is not None}
                if DataValidator.CATEGORY_COLUMNS.issubset(headers):
                    wb.close()
                    return "Kategorie"

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = {str(cell.value) for cell in sheet[1] if cell.value is not None}
                if DataValidator.CATEGORY_COLUMNS.issubset(headers):
                    wb.close()
                    return sheet_name

            wb.close()
            default = wb.sheetnames[0] if wb.sheetnames else "Sheet1"
            return default
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {file_path}")
        except PermissionError:
            raise PermissionError(f"Permission denied: {file_path}")
        except Exception as e:
            raise RuntimeError(f"An error occurred: {e}")

    @staticmethod
    def get_data_summary(df: pd.DataFrame, data_type: str = "sales") -> dict[str, Any]:
        summary: dict[str, Any] = {
            "rows": len(df),
            "columns": len(df.columns),
            "column_names": df.columns.tolist(),
        }

        if data_type == "sales":
            if "sku" in df.columns:
                summary["unique_skus"] = df["sku"].nunique()
            if "order_id" in df.columns:
                summary["unique_orders"] = df["order_id"].nunique()
            if "data" in df.columns and len(df) > 0:
                min_date = str(df["data"].min())
                max_date = str(df["data"].max())
                summary["date_range"] = f"{min_date} to {max_date}"

        elif data_type == "stock":
            if "sku" in df.columns:
                summary["unique_skus"] = df["sku"].nunique()
            if "aktywny" in df.columns:
                summary["active_skus"] = (df["aktywny"] == 1).sum()
                summary["inactive_skus"] = (df["aktywny"] == 0).sum()

        return summary
